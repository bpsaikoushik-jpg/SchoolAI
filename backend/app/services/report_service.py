"""Reports module.

Deliberately contains ZERO new business-logic queries. Every number in
every report is produced by calling the same services the Founder and
Principal portals already call (FounderService, FounderAnalyticsService,
PrincipalService, PrincipalAnalyticsService). This module's only job is
to (1) shape those results into a uniform, tabular ReportBundle and
(2) render that bundle as PDF / Excel / CSV.

Scope rules mirror every other dual-scope module in this codebase
(compare FounderAnalyticsService vs PrincipalAnalyticsService):
- FOUNDER / ADMIN  -> organization-wide, aggregated across every active
  school (loops the active schools exactly like FounderAnalyticsService
  already does internally).
- PRINCIPAL        -> scoped to the caller's own school_id only.
"""
import csv
import io
import logging
from datetime import datetime, timezone
from typing import List

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.school import School
from app.models.user import User, UserRole
from app.schemas.reports import ReportType
from app.services.founder_service import FounderService
from app.services.founder_analytics_service import FounderAnalyticsService
from app.services.principal_service import PrincipalService
from app.services.principal_analytics_service import PrincipalAnalyticsService

logger = logging.getLogger("schoolai.reports")

REPORT_CATALOG = {
    ReportType.STUDENTS: ("Students Report", "Roster of every student, class, and grade level."),
    ReportType.TEACHERS: ("Teachers Report", "Roster of every teacher, subjects taught, and performance."),
    ReportType.SCHOOLS: ("Schools Report", "Every school with headline performance figures."),
    ReportType.ATTENDANCE: ("Attendance Report", "Present / absent / late breakdown."),
    ReportType.HOMEWORK: ("Homework Report", "Homework assigned vs. submitted, per class."),
    ReportType.ASSIGNMENTS: ("Assignments Report", "Homework assigned vs. submitted, per class."),
    ReportType.EXAMS: ("Exams Report", "Every exam with results recorded and average score."),
    ReportType.RESULTS: ("Results Report", "Every exam with results recorded and average score."),
    ReportType.ORGANIZATION: ("Organization Analytics Report", "Platform-wide headline figures and role breakdown."),
}

# Report types that only make sense at the organization level.
FOUNDER_ONLY_TYPES = {ReportType.ORGANIZATION}


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.founder_service = FounderService(db)
        self.founder_analytics = FounderAnalyticsService(db)
        self.principal_service = PrincipalService(db)
        self.principal_analytics = PrincipalAnalyticsService(db)

    # -----------------------------------------------------------------
    # Public entry points
    # -----------------------------------------------------------------
    def available_report_types(self, user: User) -> List[dict]:
        types = list(REPORT_CATALOG.items())
        if user.role == UserRole.PRINCIPAL:
            types = [(t, meta) for t, meta in types if t not in FOUNDER_ONLY_TYPES]
        return [
            {"value": t.value, "label": meta[0], "description": meta[1]}
            for t, meta in types
        ]

    async def build_report(self, report_type: ReportType, user: User) -> dict:
        if user.role == UserRole.PRINCIPAL and report_type in FOUNDER_ONLY_TYPES:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="This report is only available at the organization (Founder) level",
            )

        is_org_wide = user.role in (UserRole.FOUNDER, UserRole.ADMIN)
        scope = "Organization (all schools)" if is_org_wide else await self._school_name(user.school_id)

        builder = self._builders()[report_type]
        columns, rows, summary = await builder(is_org_wide, user)

        title, _ = REPORT_CATALOG[report_type]
        return {
            "report_type": report_type.value,
            "title": title,
            "scope": scope,
            "generated_at": datetime.now(timezone.utc),
            "summary": summary,
            "columns": columns,
            "rows": rows,
        }

    def _builders(self):
        return {
            ReportType.STUDENTS: self._build_students,
            ReportType.TEACHERS: self._build_teachers,
            ReportType.SCHOOLS: self._build_schools,
            ReportType.ATTENDANCE: self._build_attendance,
            ReportType.HOMEWORK: self._build_homework,
            ReportType.ASSIGNMENTS: self._build_homework,
            ReportType.EXAMS: self._build_exams,
            ReportType.RESULTS: self._build_exams,
            ReportType.ORGANIZATION: self._build_organization,
        }

    async def _school_name(self, school_id) -> str:
        if not school_id:
            return "No school assigned"
        school = (await self.db.execute(select(School).where(School.id == school_id))).scalars().first()
        return school.name if school else "Unknown school"

    async def _active_schools(self) -> List[School]:
        return (await self.db.execute(
            select(School).where(School.deleted_at.is_(None)).order_by(School.name)
        )).scalars().all()

    # -----------------------------------------------------------------
    # Students
    # -----------------------------------------------------------------
    async def _build_students(self, is_org_wide: bool, user: User):
        columns = [
            {"key": "full_name", "label": "Name"},
            {"key": "email", "label": "Email"},
            {"key": "student_id_number", "label": "Student ID"},
            {"key": "grade_level", "label": "Grade"},
            {"key": "class_name", "label": "Class"},
            {"key": "is_active", "label": "Active"},
        ]
        if is_org_wide:
            columns.insert(1, {"key": "school_name", "label": "School"})

        rows: List[dict] = []
        if is_org_wide:
            for school in await self._active_schools():
                students = await self.principal_service.list_students(school.id, limit=1000)
                for s in students:
                    rows.append({**s, "school_name": school.name})
        else:
            rows = await self.principal_service.list_students(user.school_id, limit=1000)

        summary = [{"label": "Total students", "value": len(rows)}]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Teachers
    # -----------------------------------------------------------------
    async def _build_teachers(self, is_org_wide: bool, user: User):
        columns = [
            {"key": "full_name", "label": "Name"},
            {"key": "email", "label": "Email"},
            {"key": "employee_id", "label": "Employee ID"},
            {"key": "specialization", "label": "Specialization"},
            {"key": "subjects_count", "label": "Subjects"},
            {"key": "is_active", "label": "Active"},
        ]
        if is_org_wide:
            columns.insert(1, {"key": "school_name", "label": "School"})

        rows: List[dict] = []
        if is_org_wide:
            for school in await self._active_schools():
                teachers = await self.principal_service.list_teachers(school.id, limit=1000)
                for t in teachers:
                    rows.append({**t, "school_name": school.name})
        else:
            rows = await self.principal_service.list_teachers(user.school_id, limit=1000)

        summary = [{"label": "Total teachers", "value": len(rows)}]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Schools
    # -----------------------------------------------------------------
    async def _build_schools(self, is_org_wide: bool, user: User):
        columns = [
            {"key": "school_name", "label": "School"},
            {"key": "total_students", "label": "Students"},
            {"key": "total_teachers", "label": "Teachers"},
            {"key": "total_classes", "label": "Classes"},
            {"key": "average_score", "label": "Avg. Score"},
        ]
        if is_org_wide:
            rows = await self.founder_analytics.school_performance()
        else:
            rows = [await self.principal_analytics.school_performance(user.school_id)]

        summary = [{"label": "Schools", "value": len(rows)}]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Attendance
    # -----------------------------------------------------------------
    async def _build_attendance(self, is_org_wide: bool, user: User):
        if is_org_wide:
            columns = [
                {"key": "school_name", "label": "School"},
                {"key": "present", "label": "Present"},
                {"key": "absent", "label": "Absent"},
                {"key": "late", "label": "Late"},
                {"key": "total_marked", "label": "Total Marked"},
            ]
            overview = await self.founder_analytics.attendance_overview()
            rows = overview["by_school"]
            summary = [
                {"label": "Date", "value": str(overview["date"])},
                {"label": "Present", "value": overview["present"]},
                {"label": "Absent", "value": overview["absent"]},
                {"label": "Late", "value": overview["late"]},
            ]
        else:
            columns = [
                {"key": "class_name", "label": "Class"},
                {"key": "present", "label": "Present"},
                {"key": "absent", "label": "Absent"},
                {"key": "late", "label": "Late"},
                {"key": "total_marked", "label": "Total Marked"},
            ]
            overview = await self.principal_service.attendance_overview(user.school_id)
            rows = overview["by_class"]
            summary = [
                {"label": "Date", "value": str(overview["date"])},
                {"label": "Present", "value": overview["present"]},
                {"label": "Absent", "value": overview["absent"]},
                {"label": "Late", "value": overview["late"]},
            ]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Homework / Assignments
    # -----------------------------------------------------------------
    async def _build_homework(self, is_org_wide: bool, user: User):
        if is_org_wide:
            columns = [
                {"key": "school_name", "label": "School"},
                {"key": "total_assigned", "label": "Assigned"},
                {"key": "total_submitted", "label": "Submitted"},
                {"key": "total_students_covered", "label": "Students Covered"},
                {"key": "completion_rate", "label": "Completion %"},
            ]
            rows = await self.founder_analytics.homework_overview()
        else:
            columns = [
                {"key": "title", "label": "Title"},
                {"key": "class_name", "label": "Class"},
                {"key": "due_date", "label": "Due Date"},
                {"key": "submitted", "label": "Submitted"},
                {"key": "total_students", "label": "Total Students"},
                {"key": "completion_rate", "label": "Completion %"},
            ]
            rows = await self.principal_service.homework_overview(user.school_id)

        summary = [{"label": "Records", "value": len(rows)}]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Exams / Results
    # -----------------------------------------------------------------
    async def _build_exams(self, is_org_wide: bool, user: User):
        if is_org_wide:
            columns = [
                {"key": "school_name", "label": "School"},
                {"key": "total_exams", "label": "Exams"},
                {"key": "results_recorded", "label": "Results Recorded"},
                {"key": "average_score", "label": "Avg. Score"},
            ]
            rows = await self.founder_analytics.exam_overview()
        else:
            columns = [
                {"key": "title", "label": "Exam"},
                {"key": "subject_name", "label": "Subject"},
                {"key": "class_name", "label": "Class"},
                {"key": "date", "label": "Date"},
                {"key": "results_recorded", "label": "Results Recorded"},
                {"key": "average_score", "label": "Avg. Score"},
            ]
            rows = await self.principal_service.exams_overview(user.school_id)

        summary = [{"label": "Records", "value": len(rows)}]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Organization analytics (Founder/Admin only)
    # -----------------------------------------------------------------
    async def _build_organization(self, is_org_wide: bool, user: User):
        columns = [
            {"key": "role", "label": "Role"},
            {"key": "count", "label": "Count"},
        ]
        overview = await self.founder_analytics.org_overview()
        rows = await self.founder_analytics.role_breakdown()
        summary = [
            {"label": "Total schools", "value": overview["total_schools"]},
            {"label": "Active schools", "value": overview["active_schools"]},
            {"label": "Total students", "value": overview["total_students"]},
            {"label": "Total teachers", "value": overview["total_teachers"]},
            {"label": "Total parents", "value": overview["total_parents"]},
            {"label": "Average score", "value": overview["average_score"]},
        ]
        return columns, rows, summary

    # -----------------------------------------------------------------
    # Exporters
    # -----------------------------------------------------------------
    @staticmethod
    def _cell(value) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        return str(value)

    def to_csv(self, bundle: dict) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([c["label"] for c in bundle["columns"]])
        for row in bundle["rows"]:
            writer.writerow([self._cell(row.get(c["key"])) for c in bundle["columns"]])
        return buffer.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 CSV cleanly

    def to_excel(self, bundle: dict) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = bundle["title"][:31]  # Excel sheet name limit

        ws.append([bundle["title"]])
        ws.append([f"Scope: {bundle['scope']}"])
        ws.append([f"Generated: {bundle['generated_at'].strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])
        if bundle["summary"]:
            for item in bundle["summary"]:
                ws.append([item["label"], self._cell(item["value"])])
            ws.append([])

        header_row_idx = ws.max_row + 1
        ws.append([c["label"] for c in bundle["columns"]])
        for cell in ws[header_row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        for row in bundle["rows"]:
            ws.append([self._cell(row.get(c["key"])) for c in bundle["columns"]])

        for idx, col in enumerate(bundle["columns"], start=1):
            width = max(len(col["label"]), 12)
            ws.column_dimensions[get_column_letter(idx)].width = min(width + 6, 40)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def to_pdf(self, bundle: dict) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        out = io.BytesIO()
        doc = SimpleDocTemplate(
            out, pagesize=landscape(A4),
            leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(bundle["title"], styles["Title"]),
            Paragraph(f"Scope: {bundle['scope']}", styles["Normal"]),
            Paragraph(f"Generated: {bundle['generated_at'].strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
            Spacer(1, 0.4 * cm),
        ]

        if bundle["summary"]:
            summary_data = [[item["label"], self._cell(item["value"])] for item in bundle["summary"]]
            summary_table = Table(summary_data, hAlign="LEFT")
            summary_table.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1E3A5F")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5 * cm))

        header = [c["label"] for c in bundle["columns"]]
        data = [header] + [
            [self._cell(row.get(c["key"])) for c in bundle["columns"]] for row in bundle["rows"]
        ]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)

        if not bundle["rows"]:
            elements.append(Spacer(1, 0.5 * cm))
            elements.append(Paragraph("No records found for this report.", styles["Normal"]))

        doc.build(elements)
        return out.getvalue()
